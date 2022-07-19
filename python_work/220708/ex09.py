from openpyxl import load_workbook
from yaml import load

wb = load_workbook('c.xlsx')
ws = wb.active

ws['A11'] = "=SUM(A1;A10)"
#한줄 삽입
ws.append([11,22,33,44,55,66,77,88,99])
ws.append([])

# for cell in ws.iter_rows:
#     for

wb.save('c.xlsx')
wb.close()
