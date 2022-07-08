from openpyxl import Workbook

wb = Workbook()
ws = wb.active  # 워크시트 활성화

ws['A1'] = '123123'
ws['A2'] = '안녕하세요'
ws.title = "처음에 자동 sheet"
ws.sheet_properties.tabColor = "1072BA"

ms1 = wb.create_sheet("MySheet")
ms1['B1'] = "안녕 친구야~"
ms1['B2'] = "하하하..."


ms2 = wb.create_sheet("MySheet",0)
ms2['C1'] = "그래 안녕~"
ms2['C2'] = "반가워!..."


ms3 = wb["MySheet"]

wb.save('a.xlsx')
wb.close()